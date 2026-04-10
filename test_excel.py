import sys
try:
    import openpyxl
    wb = openpyxl.load_workbook('C:/s/Sistema/wiao/wioa/WIOA/Tabla - AMSI - Original - 5enero2026.xlsx', data_only=True, read_only=True)
    sheet = wb['Grupo 1 – 2026']
    
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = row
    
    # Print the columns indices and names:
    for i, col_name in enumerate(headers):
        if col_name:
            # Excel columns: A = 0, B = 1, C = 2...
            # A=0, B=1, C=2, D=3, E=4, F=5, G=6, 'H'=7
            # Let's print the letter as well for easy mapping
            col_letter = openpyxl.utils.get_column_letter(i+1)
            print(f"Col {col_letter}: {col_name}")

except Exception as e:
    print("Error:", e)
