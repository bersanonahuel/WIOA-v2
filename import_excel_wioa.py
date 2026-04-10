import os
import sys
import django
import datetime
import argparse
import re
import openpyxl
from django.db import transaction

# Definimos el proyecto django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "WIOA.settings")
django.setup()

from django.contrib.auth.models import User
from apps.administracion.models import Alumno, Maestro, Cargo, Escuela, Municipio
from apps.proyecto.models import Proyecto, Servicio, ServiciosProyecto
from apps.registros.models import Registro, RegistroDetalle

def parse_start_time(time_str):
    # time_str ej: "4 pm a 6 pm", "8:30 am a 11:30 am", "12:00 pm a 2:00 pm"
    if not time_str:
        return datetime.time(8, 0) # 8:00 AM fallback
    
    time_str = str(time_str).lower().strip()
    # capturar la primera parte
    match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)', time_str)
    if match:
        hours = int(match.group(1))
        mins = int(match.group(2)) if match.group(2) else 0
        ampm = match.group(3)
        
        if ampm == 'pm' and hours != 12:
            hours += 12
        elif ampm == 'am' and hours == 12:
            hours = 0
            
        return datetime.time(hours, mins)
    return datetime.time(8, 0)

def split_name(full_name):
    # Separar inteligentemente "Alex D Medina Guzmán" -> Nombre: "Alex D", Paterno: "Medina", Materno: "Guzmán"
    parts = full_name.strip().split()
    if len(parts) == 0:
        return "Desconocido", "", ""
    elif len(parts) == 1:
        return parts[0], "", ""
    elif len(parts) == 2:
        return parts[0], parts[1], ""
    elif len(parts) == 3:
        return parts[0] + " " + parts[1], parts[2], ""
    else:
        # Asume "Nombre1 Nombre2 Apellido1 Apellido2"
        # 4 partes: 2 para nombre, 1 paterno, 1 materno. "Juan Carlos Perez Gomez" -> "Juan Carlos", "Perez", "Gomez"
        nombre = " ".join(parts[:-2])
        paterno = parts[-2]
        materno = parts[-1]
        return nombre, paterno, materno

def create_or_get_generic_objects(admin_user):
    print("[*] Creando entidades genéricas...")
    muni, _ = Municipio.objects.get_or_create(nombre='Genérico')
    escuela, _ = Escuela.objects.get_or_create(nombre='Escuela Base General', defaults={'municipio': muni})
    
    cargo_tutor, _ = Cargo.objects.get_or_create(nombre='Tutor')
    cargo_mentor, _ = Cargo.objects.get_or_create(nombre='Mentor')
    cargo_consejero, _ = Cargo.objects.get_or_create(nombre='Consejero')
    
    serv_tut, _ = Servicio.objects.get_or_create(nombre='Tutorías')
    serv_ment, _ = Servicio.objects.get_or_create(nombre='Mentoría')
    serv_cons, _ = Servicio.objects.get_or_create(nombre='Consejería Abarcadora')
    
    proyecto, _ = Proyecto.objects.get_or_create(
        nombre="Proyecto AMSI - Grupos 2026",
    )
    
    # Crear ServiciosProyecto
    sp_tut, _ = ServiciosProyecto.objects.get_or_create(
        proyecto=proyecto, servicio=serv_tut, defaults={
            'precio_por_hora_participante': 15.00, # Valores defaults, a ajustar si es necesario
            'cantidad_participantes': 100,
            'total_horas': 56,
            'tipo_facturacion': 'Por hora'
        }
    )
    sp_ment, _ = ServiciosProyecto.objects.get_or_create(
        proyecto=proyecto, servicio=serv_ment, defaults={
            'precio_por_hora_participante': 15.00,
            'cantidad_participantes': 100,
            'total_horas': 48,
            'tipo_facturacion': 'Por hora'
        }
    )
    sp_cons, _ = ServiciosProyecto.objects.get_or_create(
        proyecto=proyecto, servicio=serv_cons, defaults={
            'precio_por_hora_participante': 15.00,
            'cantidad_participantes': 100,
            'total_horas': 20,
            'tipo_facturacion': 'Por hora'
        }
    )
    
    return {
        'escuela': escuela,
        'cargos': {
            'Tutor': cargo_tutor,
            'Mentor': cargo_mentor,
            'Consejero': cargo_consejero
        },
        'servicios_proyecto': {
            'Tutor': sp_tut,
            'Mentor': sp_ment,
            'Consejero': sp_cons
        },
        'proyecto': proyecto
    }

def process_profesional(name, email, cargo_obj):
    if not name or str(name).strip() == "":
        return None
    
    name = str(name).strip()
    email = str(email).strip() if email else f"{name.replace(' ', '').lower()}@generico.com"
    
    nombre, paterno, materno = split_name(name)
    apellido_final = (paterno + " " + materno).strip()
    if not apellido_final:
        apellido_final = "S/A"
        
    usr, created_usr = User.objects.get_or_create(username=email, defaults={
        'email': email,
        'first_name': nombre[:30],
        'last_name': apellido_final[:30]
    })
    if created_usr:
        usr.set_password('Mandalay2026*')
        usr.save()
        
    maestro, _ = Maestro.objects.get_or_create(
        usuario=usr, 
        defaults={
            'nombre': nombre,
            'apellido': apellido_final,
            'cargo': cargo_obj
        }
    )
    return maestro
    

def handle_excel_import(dry_run=True, sheet_target="Grupo 1 - 2026"):
    print(f"[{'DRY-RUN' if dry_run else 'ACTIVE'}] Arrancando importación...")
    
    if not User.objects.filter(is_superuser=True).exists():
        admin = User.objects.create_superuser('admin_db', 'admin@example.com', 'admin')
    else:
        admin = User.objects.filter(is_superuser=True).first()

    base_objects = create_or_get_generic_objects(admin)
    escuela = base_objects['escuela']
    cargos = base_objects['cargos']
    sps = base_objects['servicios_proyecto']
    proyecto = base_objects['proyecto']

    wb_path = os.path.join(os.path.dirname(__file__), 'WIOA', 'Tabla - AMSI - Original - 5enero2026.xlsx')
    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    
    if sheet_target not in wb.sheetnames:
        print(f"[ERROR] La hoja '{sheet_target}' no existe en el archivo.")
        return
        
    sheet = wb[sheet_target]
    
    # Extraemos filas a una lista por simplicidad
    rows = list(sheet.iter_rows(values_only=True))
    
    # Comenzamos iteracion ignorando primera fila (headers) e iterando de 2 en 2
    student_count = 0
    records_added = 0
    
    for i in range(2, len(rows), 2): # empieza en row 2 (index 2) asumiendo row 0 y 1 son headers
        row_hours = rows[i]
        if i + 1 >= len(rows):
            break
        row_dates = rows[i+1]
        
        # Validar si hay un estudiante
        student_name = row_hours[5] # Col F
        if not student_name or str(student_name).strip() == "" or str(student_name).strip() == "None":
            continue
            
        print(f"\n--> Parseando Estudiante: {student_name}")
        student_count += 1
        
        # Alumno
        nombre, paterno, materno = split_name(str(student_name))
        email_al = row_hours[4]  # Col E
        celular_al = row_hours[3] # Col D
        
        alumno, created_al = Alumno.objects.get_or_create(
            nombre=nombre,
            apellidoPaterno=paterno,
            apellidoMaterno=materno,
            defaults={
                'nivelEscolar': 'No Especifico',
                'email': email_al if email_al else '',
                'telefono': celular_al if celular_al else '',
                'escuela': escuela
            }
        )
        if created_al:
            print(f"    [NUEVO] Alumno {alumno}")
            
        proyecto.alumnos.add(alumno)
            
        # Profesionales
        # Fase I: Tutorías (Col G=6, Col H=7, Col I=8)
        # Horas: indices 9 al 36
        nombre_tutor = row_hours[6]
        email_tutor = row_hours[7]
        horario_tutor = row_hours[8]
        tutor = process_profesional(nombre_tutor, email_tutor, cargos['Tutor'])
        if tutor:
            tutor.alumnos.add(alumno)
            proyecto.maestros.add(tutor)
            
        # Fase II: Mentorías (Típicamente en col 38, 39, 40)
        nombre_mentor = row_hours[38] if len(row_hours) > 38 else None
        email_mentor = row_hours[39] if len(row_hours) > 39 else None
        horario_mentor = row_hours[40] if len(row_hours) > 40 else None
        mentor = process_profesional(nombre_mentor, email_mentor, cargos['Mentor'])
        if mentor:
            mentor.alumnos.add(alumno)
            proyecto.maestros.add(mentor)
            
        # Fase III: Consejería (Típicamente en col 66, 67, 68) pero ajustamos según el array json que vimos que Mentoría era 42-65 (24 elems) -> Total 65 -> Consejero es 66? 
        # Revisando indices reales de info2.json: Mentoría termina en col 64, 65 es total. Consejero es 66: "Keila Santos Díaz"
        nombre_cons = row_hours[66] if len(row_hours) > 66 else None
        email_cons = row_hours[67] if len(row_hours) > 67 else None
        horario_cons = row_hours[68] if len(row_hours) > 68 else None
        consejero = process_profesional(nombre_cons, email_cons, cargos['Consejero'])
        if consejero:
            consejero.alumnos.add(alumno)
            proyecto.maestros.add(consejero)
            
        # === INSERCION DE HORAS ===
        fases = [
            ('Tutor', tutor, horario_tutor, sps['Tutor'], 9, 36),
            ('Mentor', mentor, horario_mentor, sps['Mentor'], 41, 64),
            ('Consejero', consejero, horario_cons, sps['Consejero'], 69, 72)
        ]
        
        for cargo_nombre, maestro_obj, horario_texto, serv_proy, start_idx, end_idx in fases:
            if not maestro_obj:
                continue
                
            start_time = parse_start_time(horario_texto)
            
            # Busco/Creo el Registro base (Agrupador)
            registro, _ = Registro.objects.get_or_create(
                proyecto_servicio=serv_proy,
                alumno=alumno,
                usuario=admin
            )
            
            last_dt_val = None
            empty_dt_count = 0
            for col_idx in range(start_idx, end_idx + 1):
                if col_idx >= len(row_hours):
                    break
                    
                hs = row_hours[col_idx]
                dt = row_dates[col_idx]
                
                try:
                    horas_float = float(hs) if hs is not None and hs != "" else 0
                except ValueError:
                    horas_float = 0
                    
                if dt:
                    empty_dt_count = 0
                    # Parsear datetime y guardarlo como el último válido
                    if isinstance(dt, datetime.datetime):
                        last_dt_val = dt.date()
                    elif isinstance(dt, datetime.date):
                        last_dt_val = dt
                    else:
                        try:
                            last_dt_val = datetime.datetime.strptime(str(dt).split(' ')[0], "%Y-%m-%d").date()
                        except:
                            pass
                else:
                    if last_dt_val:
                        empty_dt_count += 1
                            
                if horas_float > 0 and last_dt_val:
                    current_dt_val = last_dt_val + datetime.timedelta(days=empty_dt_count)
                    inicio_dt = datetime.datetime.combine(current_dt_val, start_time)
                    fin_dt = inicio_dt + datetime.timedelta(hours=horas_float)
                    
                    # Evitar duplicados
                    if not RegistroDetalle.objects.filter(registro=registro, fechaHoraInicio=inicio_dt).exists():
                        RegistroDetalle.objects.create(
                            fechaHoraInicio=inicio_dt,
                            fechaHoraFin=fin_dt,
                            registro=registro,
                            usuario=admin, # O usar maestro_obj.usuario
                            comentario=f"Sesión importada de Excel ({cargo_nombre})"
                        )
                        records_added += 1

    print(f"\n[OK] Analizados {student_count} estudiantes.")
    print(f"[OK] Se insertarían/insertaron {records_added} sesiones (RegistroDetalle).")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Importador de Excel a Django Models')
    parser.add_argument('--commit', action='store_true', help='Ejecuta los cambios en la DB (por defecto es dry-run)')
    args = parser.parse_args()
    
    try:
        with transaction.atomic():
            handle_excel_import(dry_run=not args.commit)
            if not args.commit:
                print("\n[INFO] DRY RUN finalizado con éxito. Se revertirán los cambios. Ejecute con --commit para guardar localmente.")
                raise Exception("DRY RUN ROLLBACK")
            else:
                print("\n[OK] COMMIT finalizado con éxito. Cambios guardados en base de datos.")
    except Exception as e:
        if str(e) == "DRY RUN ROLLBACK":
            pass # Es normal, rollback esperado
        else:
            print("[ERROR FATAL] La transacción falló. Detalle:", e)
            import traceback
            traceback.print_exc()
