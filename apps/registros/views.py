from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, DeleteView, UpdateView, TemplateView
from django.views.generic.edit import FormView
from django.views.generic.base import View
from django.urls.base import reverse_lazy
#from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.forms import formset_factory
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import FloatField, DecimalField, ExpressionWrapper, Count, Sum, F, Q
from .models import *
from .forms import RegistroForm, RegistroDetalleForm, FacturaForm
from apps.proyecto.models import ServiciosProyecto, Proyecto, Servicio
from apps.administracion import models
from django.db import models as django_models
 
from datetime import datetime, timedelta
from decimal import Decimal
#Para PDF
from io import BytesIO # nos ayuda a convertir un html en pdf
import os
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
from weasyprint import HTML
from weasyprint import CSS
#Filtro
from .filters import RegistroFilter, FacturaFilter
#Excel
from openpyxl import Workbook
from openpyxl.styles import Color, Fill ,PatternFill

CREAR_REGISTRO_FILE  = 'registros/registro/crearRegistro.html'
LISTAR_REGISTRO_FILE = 'registros/registro/listarRegistro.html'
LISTAR_REGISTRO_PROYECTO_FILE = 'registros/registro/listarRegistroPorProyecto.html'
CREAR_REGISTRO_DETALLE_FILE = 'registros/registro/crearRegistroDetalle.html'
LISTAR_FACTURA_FILE  = 'registros/factura/listarFactura.html'
CREAR_FACTURA_FILE   = 'registros/factura/crearFactura.html'
CREAR_REGISTRO_DETALLE_MASIVO_FILE = 'registros/registro/crearRegistroDetalleMasivo.html'

class CrearRegistro(CreateView):
    models = Registro
    form_class = RegistroForm
    template_name = CREAR_REGISTRO_FILE
    
    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {}
            try:
                action = request.POST.get('action')
                
                if action == 'getAlumnosDelProyecto':
                    proyectoId = 0
                    serviciosProyectoSelectId = request.POST.get('servicioProyectoId')
                    serviciosProyecto = None
                    if serviciosProyectoSelectId:
                        serviciosProyecto = ServiciosProyecto.objects.get(id=serviciosProyectoSelectId)
                        proyectoId = serviciosProyecto.proyecto.id
                    elif request.POST.get('proyectoId'):
                        proyectoId = request.POST.get('proyectoId')
                    
                    if proyectoId > 0:
                        p = Proyecto.objects.get(id=proyectoId)
                        
                        #Buscar alumnos del proyecto y servicio que todavia no tienen registro creado.
                        if serviciosProyectoSelectId:
                            alumnosConRegistroCreado = Registro.objects.filter(proyecto_servicio=serviciosProyectoSelectId).values_list('alumno_id', flat=True)

                        alumnos = p.alumnos.exclude(pk__in=alumnosConRegistroCreado).values('id', 'nombre', 'apellidoPaterno', 'apellidoMaterno')

                    else:
                        data['error'] = 'No se seleccionó ningún Proyecto'

                    if alumnos:
                        data = { 'alumnos': list(alumnos) }
                    else:
                        data['error'] = 'No se encontró ningún alumno asignado a este Proyecto.'
                else:
                    data['error'] = 'Ha ocurrido un error'

            except Exception as e:
                data['error'] = str(e)
            return JsonResponse(data, safe=False)
        
        if request.method == 'POST':
            form = RegistroForm(request.POST)
            registro = form.save(commit=False)
            registro.usuario = request.user
            registro.save()
            
            self.kwargs['registropk'] = registro.id

            return super().form_valid(form)

    def get_context_data(self, *args, **kwargs):
        serviciosProyecto = ServiciosProyecto.get_servicios_proyectos_del_usuario(self.request.user)

        kwargs['titulo'] = 'Crear registro'
        kwargs['serviciosProyecto'] = serviciosProyecto

        return  super(CrearRegistro,self).get_context_data(**kwargs)
    
    def get_success_url(self):
        return reverse_lazy('registros:crearRegistroDetalle',args=[self.kwargs['registropk']])

class ListarRegistro(ListView):
    model = Registro
    template_name = LISTAR_REGISTRO_FILE

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        if request.user.is_staff:
            registro_filter = RegistroFilter(request.GET, queryset=Registro.objects.all(), request=self.request)
        else:
            registro_filter = RegistroFilter(request.GET, queryset=Registro.objects.filter(usuario=self.request.user), request=self.request)
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {}
            try:
                action = request.GET.get('action')
          
                if action == 'getAlumnosDelProyecto':

                    data = get_alumnos_del_proyecto(request.GET.get('servicioProyectoId'), request.GET.get('proyectoId'), data)

                else:
                    data['error'] = 'Ha ocurrido un error'

            except Exception as e:
                data['error'] = str(e)
            return JsonResponse(data, safe=False)
        
        return render(request, LISTAR_REGISTRO_FILE, {'filter': registro_filter})

from django.db.models import Count
class ListarRegistroPorProyecto(ListView):
    model = Registro
    template_name = LISTAR_REGISTRO_PROYECTO_FILE

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        if request.user.is_staff:
            registro_filter = RegistroFilter(request.GET, queryset=Registro.objects.all(), request=self.request)
        else:
            registro_filter = RegistroFilter(request.GET, queryset=Registro.objects.filter(usuario=self.request.user), request=self.request)
        
        registrosInstance = Registro.objects.filter(id__in=registro_filter.qs)
        
        reg_count = registrosInstance.values('proyecto_servicio__id', 'proyecto_servicio__proyecto__nombre', 'proyecto_servicio__servicio__nombre', 'proyecto_servicio__cantidad_participantes').order_by('proyecto_servicio').annotate(cant_con_registro=Count('id'))
        
        # Cantidad de registros / participantes por proyecto y servicio. 
        i= 0
        for reg in reg_count:
            #Busco todos los registros del Proyecto/Servicio
            registros_todos = Registro.objects.filter(proyecto_servicio__id=reg['proyecto_servicio__id'])
            servidos = 0
            enProceso = 0
            for regInst in registros_todos:
                hsFaltantes = regInst.calcular_total_horas_faltantes_por_alumno()
                hsRegistradas = regInst.calcular_total_horas_registradas_por_alumno()

                if hsFaltantes == '00:00':
                    servidos = servidos + 1
                elif hsRegistradas > 0:
                    enProceso = enProceso + 1
                
            cantTotalConHs = servidos + enProceso
            reg_count[i]['cant_servidos'] = servidos 
            reg_count[i]['cant_en_proceso'] = enProceso #Este muestra todos los que tienen registro creado, tengan o no hs cargadas.. reg['cant_con_registro'] - servidos
            reg_count[i]['cant_total_con_hs'] = cantTotalConHs 
            reg_count[i]['cant_faltan_identificar'] = reg['proyecto_servicio__cantidad_participantes'] - reg['cant_total_con_hs']


            #De la cantidad que faltan reg hs, dividir en los que estan cargados en el sistema sin registro y los que faltarian cargar.
            cargados = 0
            sinCargar = 0

            ps = ServiciosProyecto.objects.get(id=reg['proyecto_servicio__id'])
            proyecto = Proyecto.objects.get(id=ps.proyecto.id)
            cargados = proyecto.alumnos.all().count()
            reg_count[i]['cant_cargados'] = cargados - cantTotalConHs
            reg_count[i]['cant_sin_cargados'] = reg['cant_faltan_identificar'] - reg_count[i]['cant_cargados']


            i = i + 1

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {}
            try:
                action = request.GET.get('action')
          
                if action == 'getAlumnosDelProyecto':
                                        
                    data = get_alumnos_del_proyecto(request.GET.get('servicioProyectoId'), request.GET.get('proyectoId'), data)

                else:
                    data['error'] = 'Ha ocurrido un error'

            except Exception as e:
                data['error'] = str(e)
            return JsonResponse(data, safe=False)
        
        return render(request, LISTAR_REGISTRO_PROYECTO_FILE, {'filter': registro_filter, 'reg_count':reg_count})

def get_alumnos_del_proyecto(servicioProyectoId, proyectoId, data):
    proyecto = 0

    if servicioProyectoId:
        serviciosProyecto = ServiciosProyecto.objects.get(id=servicioProyectoId)
        proyecto = serviciosProyecto.proyecto.id
    elif proyectoId:
        proyecto = proyectoId
    
    if int(proyecto) > 0:
        p = Proyecto.objects.get(id=proyecto)
        alumnos = p.alumnos.values('id', 'nombre', 'apellidoPaterno', 'apellidoMaterno')
    else:
        #data['error'] = 'No se seleccionó ningún Proyecto'
        alumnos = Alumno.objects.all().values('id', 'nombre', 'apellidoPaterno', 'apellidoMaterno')

    if alumnos:
        data = { 'alumnos': list(alumnos) }
    else:
        data['error'] = 'No se encontró ningún alumno asignado a este Proyecto.'
    
    return data

def calcular_horas_facturadas_proyecto(proyecto, servicio):
    sp = ServiciosProyecto.objects.filter(proyecto=proyecto, servicio=servicio).first()
        
    facturaReg = Factura.objects.filter(proyectosServicios=sp)
    
    #Busco todos los reg det de las facturas ya creadas
    seg = 0
    for fact in facturaReg:
        regDet = RegistroDetalle.objects.filter(factura=fact, registro__proyecto_servicio__servicio=servicio)

        for r in regDet:
            seg = seg + r.calcular_total_hs_segundos_detalle()
    
    return seg

#Busca todos las horas ya registradas de un alumno en particlar y las suma.
def calcular_horas_registradas_alumno(registroInstance):
    registrosDetalle = RegistroDetalle.objects.filter(registro=registroInstance)
    
    seg = 0
    for det in registrosDetalle:
        seg = seg + det.calcular_total_hs_segundos_detalle()
    
    return seg


class CrearRegistroDetalle(CreateView):
    models = RegistroDetalle
    form_class = RegistroDetalleForm
    template_name = CREAR_REGISTRO_DETALLE_FILE

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            request.POST._mutable = True
            
            #Pruebas convertir time AM/PM
            format = '%Y-%m-%d %H:%M'
            date_string_inicio = request.POST['fechaHoraInicio']  #'2009-11-29 03:17:00.0000'
            inicio = datetime.strptime(date_string_inicio, format)
    
            date_string_fin = request.POST['fechaHoraFin']
            fin = datetime.strptime(date_string_fin, format)

            registroInstance = Registro.objects.get(id=self.kwargs['registropk'])
            form = RegistroDetalleForm()
            regDet = form.save(commit=False)
            regDet.fechaHoraInicio = inicio.strftime(format)
            regDet.fechaHoraFin = fin.strftime(format)
            regDet.registro = registroInstance
            regDet.usuario = request.user
            regDet.comentario = request.POST['comentario']

            proyecto = registroInstance.proyecto_servicio.proyecto
            servicio = registroInstance.proyecto_servicio.servicio
            sp = ServiciosProyecto.objects.filter(proyecto=proyecto, servicio=servicio).first()
            totalHorasSP = (sp.total_horas * 3600)
            
            hsRegistradasAlumno = calcular_horas_registradas_alumno(registroInstance)
            
            #Validar que no se exceda del total de Hs del proyecto.

            timediff = (fin - inicio)
            hsRegistroActual = timediff.seconds
            
            if sp.total_horas > 0 and (hsRegistradasAlumno + hsRegistroActual) > totalHorasSP:
                error = 'Las horas que quiere registrar para el participante se exceden del Total de Horas ('+str(sp.total_horas)+') permitidas para este Proyecto y Servicio.'
                messages.error(self.request, error)
                context = {
                    'titulo': 'Crear registro de horas',
                    'registro': Registro.objects.get(id=self.kwargs['registropk']),
                    'registrosDet': RegistroDetalle.objects.filter(registro=self.kwargs['registropk']),
                    'form': RegistroDetalleForm()
                }
                return render(request, CREAR_REGISTRO_DETALLE_FILE, context)
            else:
                regDet.save()
                success = 'El registro se guardó correctamente.'
                messages.success(self.request, success)
            
            return super().form_valid(regDet)
    
    def get_context_data(self, *args, **kwargs):
        kwargs['titulo'] = 'Crear registro de horas'
        kwargs['registro'] = Registro.objects.get(id=self.kwargs['registropk'])
        kwargs['registrosDet'] = RegistroDetalle.objects.filter(registro=self.kwargs['registropk'])
        
        return super(CrearRegistroDetalle,self).get_context_data(**kwargs)
    
    def form_valid(self, form):
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('registros:crearRegistroDetalle',args=[self.kwargs['registropk']])

class EliminarRegistroDetalle(DeleteView):
    model = RegistroDetalle

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            registroDetalle = self.get_object()
            registroDetalle.delete()
            res = {
                'mensaje': 'El detalle se eliminó con éxito.'
            }
            return JsonResponse(res, safe=False)
        return super().post(request, *args, **kwargs)


class CrearRegistroDetalleMasivo(CreateView):
    models = RegistroDetalle
    form_class = RegistroDetalleForm
    template_name = CREAR_REGISTRO_DETALLE_MASIVO_FILE

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {}
            try:
                action = request.POST.get('action')
                
                if action == 'getAlumnosDelProyecto':
                    proyectoId = 0
                    serviciosProyectoSelectId = request.POST.get('servicioProyectoId')
                    serviciosProyecto = None
                    if serviciosProyectoSelectId:
                        serviciosProyecto = ServiciosProyecto.objects.get(id=serviciosProyectoSelectId)
                        proyectoId = serviciosProyecto.proyecto.id
                        
                        p = Proyecto.objects.get(id=proyectoId)
                        
                        alumnos = p.alumnos.values('id', 'nombre', 'apellidoPaterno', 'apellidoMaterno')
                    
                        if alumnos:
                            data = { 'alumnos': list(alumnos) }
                        else:
                            data['error'] = 'No se encontró ningún alumno asignado a este Proyecto.'

                    else:
                        data['error'] = 'No se seleccionó ningún Proyecto.'

                else:
                    data['error'] = 'Ha ocurrido un error'

            except Exception as e:
                data['error'] = str(e)
            return JsonResponse(data, safe=False)
        
        if request.method == 'POST':
            request.POST._mutable = True
            
            #Pruebas convertir time AM/PM
            format = '%Y-%m-%d %H:%M'
            date_string_inicio = request.POST['fechaHoraInicio']  #'2009-11-29 03:17:00.0000'
            inicio = datetime.strptime(date_string_inicio, format)
    
            date_string_fin = request.POST['fechaHoraFin']
            fin = datetime.strptime(date_string_fin, format)

            sp = ServiciosProyecto.objects.get(id=request.POST['proyectoServicioSelectId'])

            for alumno in request.POST.getlist('alumnosCheck'):
        
                registroInstance = Registro.objects.filter(alumno=alumno, proyecto_servicio=sp).first()
                alumnoInstance = Alumno.objects.get(id=alumno)

                print(' ********** registroInstance ')
                print(registroInstance)

                if not registroInstance:
                        #Creo el registro nuevo del alumno
                        print(' ----- no hay registro creado para este alumno y servicio')
                        form = RegistroForm()
                        registro = form.save(commit=False)
                        registro.alumno = alumnoInstance
                        registro.proyecto_servicio = sp
                        registro.usuario = request.user
                        
                        if registro.save():
                            registroInstance = registro
                        else:
                            error = "No se pudo crear el registro de horas para el alumno " + alumnoInstance.nombre + " " + alumnoInstance.apellidoPaterno + '.'
                            messages.error(self.request, error)

                form = RegistroDetalleForm()
                regDet = form.save(commit=False)
                regDet.fechaHoraInicio = inicio.strftime(format)
                regDet.fechaHoraFin = fin.strftime(format)
                regDet.registro = registroInstance
                regDet.usuario = request.user
                regDet.comentario = request.POST['comentario']

                totalHorasSP = (sp.total_horas * 3600)
                
                hsRegistradasAlumno = calcular_horas_registradas_alumno(registroInstance)
                
                #Validar que no se exceda del total de Hs del proyecto.
                timediff = (fin - inicio)
                hsRegistroActual = timediff.seconds
                
                if sp.total_horas > 0 and (hsRegistradasAlumno + hsRegistroActual) > totalHorasSP:
                    error = 'Las horas que quiere registrar para el participante '+ alumnoInstance.nombre + ' ' + alumnoInstance.apellidoPaterno +' se exceden del Total de Horas ('+str(sp.total_horas)+') permitidas para este Proyecto y Servicio.'
                    messages.error(self.request, error)
                    context = {
                        'titulo': 'Crear registro de horas para varios alumnos',
                        'serviciosProyecto': ServiciosProyecto.objects.all()
                    }
                    return render(request, CREAR_REGISTRO_DETALLE_MASIVO_FILE, context)
                else:
                    regDet.save()
                    success = 'El registro se guardó correctamente.'
                    messages.success(self.request, success)
            
            return super().form_valid(regDet)
    
    def get_context_data(self, *args, **kwargs):
        kwargs['titulo'] = 'Crear registro de horas para varios alumnos'
        kwargs['serviciosProyecto'] = ServiciosProyecto.objects.all()

        return super(CrearRegistroDetalleMasivo,self).get_context_data(**kwargs)
    
    def form_valid(self, form):
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('registros:crearRegistroDetalleMasivo')


class CrearFactura(CreateView):
    models = Factura
    form_class = FacturaForm
    template_name = CREAR_FACTURA_FILE

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        error = None
        mensaje = None
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {}
            try:
                action = request.POST.get('action')
                
                if action == 'getServiciosDelProyecto':

                    p = Proyecto.objects.get(id=request.POST.get('proyectoId'))
                    serviciosProyecto = ServiciosProyecto.objects.filter(proyecto=p.id)
                    
                    servicioIds = serviciosProyecto.values_list('servicio_id', flat=True) #.values('id', 'nombre')
                    servicios = Servicio.objects.filter(pk__in=servicioIds).values('id', 'nombre')
                    
                    if servicios:
                        data = { 'servicios': list(servicios) }
                    else:
                        data['error'] = 'No se encontró ningún Servicio asignado a este Proyecto.'
                else:
                    data['error'] = 'Ha ocurrido un error'

            except Exception as e:
                data['error'] = str(e)
            return JsonResponse(data, safe=False)

        if request.method == 'POST':
            request.POST._mutable = True

            #Pruebas convertir time AM/PM
            format = '%Y-%m-%d'
            date_string_inicio = request.POST['fechaInicio']
            inicio = datetime.strptime(date_string_inicio, format)
            inicioDate=inicio.date()
    
            date_string_fin = request.POST['fechaFin']
            fin = datetime.strptime(date_string_fin, format)
            finDate = fin.date()
            request.POST['proyectosServicios'] = ServiciosProyecto.objects.all().first() #Poner uno por defecto para que valide el formulario.
            request.POST['usuario'] = request.user
            
            form = FacturaForm(request.POST)
            
            if form.is_valid():
                proyectoSelId = request.POST['proyectoSelId']                
                # Validar que no se exeda la cantidad de HS a Facturar establecidas del ProyectoServicio

                #Busco hs ya facturadas
                for serv in request.POST.getlist('serviciosCheck'):
                    hsAFacturarServicio = 0
                    hsFacturadasServicio = 0

                    sp = ServiciosProyecto.objects.filter(proyecto=proyectoSelId, servicio=serv).first()
                    if not sp:
                         messages.error(self.request, f'No se encontró el servicio con ID {serv} para el proyecto seleccionado.')
                         return render(request, CREAR_FACTURA_FILE, {'form':form, 'proyectos': Proyecto.objects.all()})

                    totalHorasSP = (sp.total_horas * sp.cantidad_participantes * 3600)

                    hsFacturadasServicio = calcular_horas_facturadas_proyecto(proyectoSelId, serv)
                    
                    #Sumar horas que se quieren facturar AHORA
                    seg = 0
                    detallesAFacturar = RegistroDetalle.objects.filter( fechaHoraInicio__date__gte=inicioDate, fechaHoraFin__date__lte=finDate, factura=None, registro__proyecto_servicio=sp )
                    
                    for detF in detallesAFacturar:
                        seg = seg + detF.calcular_total_hs_segundos_detalle()
                    
                    hsAFacturarServicio = seg
                    
                    if sp.total_horas > 0 and sp.cantidad_participantes > 0 and (hsFacturadasServicio + hsAFacturarServicio) > totalHorasSP:
                        error = 'Las horas que quiere facturar se exceden del Total de horas permitidas para el Proyecto "'+ sp.proyecto.nombre +'" y el Servicio "'+ sp.servicio.nombre+'".'
                        messages.error(self.request, error)
                    
                    if not detallesAFacturar:
                        mensaje = 'No se encontraron registros para facturar en el período seleccionado.'
                        messages.warning(self.request, mensaje)

                if error or mensaje:
                    return render(request, CREAR_FACTURA_FILE, {'form':form, 'proyectos': Proyecto.objects.all()})
                else:
                    factura = form.save(commit=False)
                    factura.fechaInicio = inicio.strftime(format)
                    factura.fechaFin = fin.strftime(format)
                    factura.fechaCreacion = fin.strftime(format)
                    factura.cliente = Cliente.objects.get(id=request.POST['cliente'])
                    factura.proveedor = Proveedor.objects.get(id=request.POST['proveedor'])
                    factura.usuario = request.user
                    factura.save()
                    
                    #Guardar tabla intermedia Factura-ProyectoServicios
                    for serv in request.POST.getlist('serviciosCheck'):
                        sp = ServiciosProyecto.objects.filter(proyecto=proyectoSelId, servicio=serv).first()
                        factura.proyectosServicios.add(sp)
                    
                    factura.save()
                    
                    messages.success(self.request, 'La factura se creó correctamente.')

                    #Buscar todos los registros de hs que se incluyan en el periodo ingresado.
                    detalles = RegistroDetalle.objects.filter( fechaHoraInicio__date__gte=inicioDate, fechaHoraFin__date__lte=finDate, factura=None )
                    listaProyServFactura = factura.proyectosServicios.values_list('id',flat=True)

                    if detalles:
                        for det in detalles:
                            if det.registro.proyecto_servicio.id in listaProyServFactura:
                                det.factura = factura
                                det.save()
                    
                    return render(request, LISTAR_FACTURA_FILE, {'filter':FacturaFilter(request.GET, queryset=Factura.objects.all())} )
            else:
                print('NO SE GUARDO, ERROR: ', form.errors)
                return super().form_valid(form)


    def form_valid(self, form):
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('registros:listarFactura')
    
    def get_context_data(self, **kwargs):
        kwargs['proyectos'] = Proyecto.objects.all()
        return super(CrearFactura,self).get_context_data(**kwargs)

class EditarFactura(UpdateView):
    model = Factura
    form_class = FacturaForm
    template_name = CREAR_FACTURA_FILE
    success_url = reverse_lazy('registros:listarFactura')

    def get_context_data(self, **kwargs):
        kwargs['boton'] = 'Modificar'
        kwargs['titulo'] = 'Editar Factura'
        return super(EditarFactura,self).get_context_data(**kwargs)

class ListarFactura(ListView):
    model = Factura
    template_name = LISTAR_FACTURA_FILE
    
    def get(self, request, *args, **kwargs):
        factura_filter = FacturaFilter(request.GET, queryset=Factura.objects.all())
        
        return render(request, LISTAR_FACTURA_FILE, {'filter': factura_filter})

class PrintPdf(View):
    def get(self, request, *args, **kwargs):
        template = get_template("registros/documentos/imprimirFactura.html")
        
        facturaInstance = Factura.objects.get(id=self.kwargs['pk'])
        
        registrosDetalle = dict()
        registrosDetalle = RegistroDetalle.objects.filter(factura=facturaInstance.id).order_by('usuario', 'registro', 'fechaHoraInicio')
        
        # Usar DurationField para cálculos de tiempo internos
        diffMiliseg = registrosDetalle.values('registro_id').order_by('registro_id').annotate(timesDif = ExpressionWrapper( (F("fechaHoraFin") - F("fechaHoraInicio")), output_field=django_models.DurationField()) )
        sumaTiempoPorRegistro = diffMiliseg.values('registro_id', 'registro__alumno__nombre', 'registro__alumno__apellidoPaterno', 'registro__alumno__apellidoMaterno', servicio=F('registro__proyecto_servicio__servicio__nombre'), escuela=F('registro__alumno__escuela__nombre')).order_by('registro_id').annotate(tiempo= Sum('timesDif')).order_by('registro__alumno__nombre')
        
        usuariosList = RegistroDetalle.objects.filter(factura=facturaInstance.id).values('usuario', 'usuario__first_name', 'usuario__last_name').distinct()
        
        #Total participantes Matriculados
        #totalParticipantesMatriculados = facturaInstance.proyectosServicios.all().aggregate(total=Sum('cantidad_participantes'))
        totalParticipantesMatriculados = 0
        sp_first = facturaInstance.proyectosServicios.all().first()
        if sp_first:
            totalParticipantesMatriculados = sp_first.cantidad_participantes    
        
        # ······ Detalle de Hs por MES
        registrosPorMes = dict()
        nro = 1
        tutoriaHsTotal = 0
        mentoriaHsTotal = 0
        conserjeriaHsTotal = 0
        jsHsTotal = 0
        seguimientoHsTotal = 0
        totalHorasServicios = 0
        
        for i in range(1, 13): 
            #regDetMes = RegistroDetalle.objects.filter(factura=facturaInstance.id, fechaHoraInicio__month = i).values('fechaHoraInicio__month', servicioId=F('registro__proyecto_servicio__servicio')).annotate(cant=Count('registro', distinct=True)).order_by('fechaHoraInicio__month', '-registro__proyecto_servicio__servicio')
            #cantParticipantes = ([regDet.registro_id for regDet in regDetMes])

            #Calculo cantidad de Participantes x Servicio x Mes
            regDetMes = RegistroDetalle.objects.filter(factura=facturaInstance.id, fechaHoraInicio__month = i).values('fechaHoraInicio__month').aggregate(
                tutoria=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=1)),
                mentoria=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=2)),
                conserjeria=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=3)),
                js=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=4)),
                seguimiento=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=5)),
            )
                        
            if regDetMes:
                if regDetMes['tutoria'] > 0 or regDetMes['mentoria'] > 0 or regDetMes['conserjeria'] > 0 or regDetMes['js'] > 0 or regDetMes['seguimiento'] > 0:
                    for servicio in range(1,6):

                        seg = 0
                        reg = RegistroDetalle.objects.filter(factura=facturaInstance.id, fechaHoraInicio__month = i, registro__proyecto_servicio__servicio=servicio)
                        for r in reg:
                            seg = seg + r.calcular_total_hs_segundos_detalle()


                        hs = convertir_tiempo(seg)
                        if servicio == 1:
                            regDetMes['tutoriaHs'] = hs
                            tutoriaHsTotal = tutoriaHsTotal + seg
                        if servicio == 2:
                            regDetMes['mentoriaHs'] = hs
                            mentoriaHsTotal = mentoriaHsTotal + seg
                        if servicio == 3:
                            regDetMes['conserjeriaHs'] = hs
                            conserjeriaHsTotal = conserjeriaHsTotal + seg
                        if servicio == 4:
                            regDetMes['jsHs']  = hs
                            jsHsTotal = jsHsTotal + seg
                        if servicio == 5:
                            regDetMes['seguimientoHs']  = hs
                            seguimientoHsTotal = seguimientoHsTotal + seg
                    
                    regDetMes['mes'] = i
                    registrosPorMes[nro] = regDetMes
                    nro = nro+1

                    totalHorasServicios = tutoriaHsTotal + mentoriaHsTotal + conserjeriaHsTotal + jsHsTotal + seguimientoHsTotal
        

        # TOTALES
        totTutoria = 0
        totMentoria = 0
        totConserjeria = 0
        totJs = 0
        totSeguimiento = 0

        for t in registrosPorMes.values():
            totTutoria+=t['tutoria']
            totMentoria+=t['mentoria']
            totConserjeria+=t['conserjeria']
            totJs+=t['js']
            totSeguimiento+=t['seguimiento']
        
        registrosPorMes['TOTAL'] = {
            'tutoria': totTutoria,
            'mentoria': totMentoria,
            'conserjeria': totConserjeria,
            'js': totJs,
            'seguimiento': totSeguimiento,
            'tutoriaHs': convertir_tiempo(tutoriaHsTotal),
            'mentoriaHs': convertir_tiempo(mentoriaHsTotal),
            'conserjeriaHs': convertir_tiempo(conserjeriaHsTotal),
            'jsHs': convertir_tiempo(jsHsTotal),
            'seguimientoHs': convertir_tiempo(seguimientoHsTotal)
        }

        #Cantidad de Participantes x Servicio
        regDetServicio = RegistroDetalle.objects.filter(factura=facturaInstance.id).aggregate(
            tutoriaCantPart=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=1)),
            mentoriaCantPart=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=2)),
            conserjeriaCantPart=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=3)),
            jsCantPart=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=4)),
            seguimientoCantPart=Count('registro', distinct=True, filter=Q(registro__proyecto_servicio__servicio=5)),
        )

        #totalParticipantesServidos = regDetServicio['tutoriaCantPart'] + regDetServicio['mentoriaCantPart']  + regDetServicio['conserjeriaCantPart'] + regDetServicio['jsCantPart'] + regDetServicio['seguimientoCantPart']
       

        # ······ Total y subtotal Factura
        subtotal = Decimal(0.0)
        total = Decimal(0.0)
        regDetServicio['tutoriaPrecio'] =  Decimal(0.0)
        regDetServicio['mentoriaPrecio'] =  Decimal(0.0)
        regDetServicio['conserjeriaPrecio'] =  Decimal(0.0)
        regDetServicio['jsPrecio'] =  Decimal(0.0)
        regDetServicio['seguimientoPrecio'] =  Decimal(0.0)

        for proyServ in facturaInstance.proyectosServicios.all():

            precio = 0.0
            if(proyServ.tipo_facturacion == 'Por hora'):
                precio = proyServ.precio_por_hora_participante
            else:
                precio = proyServ.calcular_precio_por_hora()

            if proyServ.servicio.id == 1:
                #regDetServicio['tutoriaPrecio'] = round(precio * regDetServicio['tutoriaCantPart'] * convertir_tiempo_decimal(tutoriaHsTotal), 2)
                regDetServicio['tutoriaPrecio'] = round(precio * convertir_tiempo_decimal(tutoriaHsTotal), 2)
            elif proyServ.servicio.id == 2:
                #regDetServicio['mentoriaPrecio'] = round(precio * regDetServicio['mentoriaCantPart'] * convertir_tiempo_decimal(mentoriaHsTotal), 2)
                regDetServicio['mentoriaPrecio'] = round(precio * convertir_tiempo_decimal(mentoriaHsTotal), 2)
            elif proyServ.servicio.id == 3:
                #regDetServicio['conserjeriaPrecio'] = round(precio * regDetServicio['conserjeriaCantPart'] * convertir_tiempo_decimal(conserjeriaHsTotal), 2)
                regDetServicio['conserjeriaPrecio'] = round(precio * convertir_tiempo_decimal(conserjeriaHsTotal), 2)
            elif proyServ.servicio.id == 4:
                #regDetServicio['jsPrecio'] = round(precio * regDetServicio['jsCantPart'] * convertir_tiempo_decimal(jsHsTotal), 2)
                regDetServicio['jsPrecio'] = round(precio * convertir_tiempo_decimal(jsHsTotal), 2)
            elif proyServ.servicio.id == 5:
                #regDetServicio['seguimientoPrecio'] = round(precio * regDetServicio['seguimientoCantPart'] * convertir_tiempo_decimal(seguimientoHsTotal), 2)
                regDetServicio['seguimientoPrecio'] = round(precio * convertir_tiempo_decimal(seguimientoHsTotal), 2)
                

        subtotal = regDetServicio['tutoriaPrecio'] + regDetServicio['mentoriaPrecio'] + regDetServicio['conserjeriaPrecio'] + regDetServicio['jsPrecio'] + regDetServicio['seguimientoPrecio']

        tax = Decimal(0.0)
        if facturaInstance.impuesto:
            tax = facturaInstance.impuesto.porcentaje
        
        precioTax = round(subtotal * (tax/100), 2)
        total = subtotal + precioTax


        context = {
            'factura': facturaInstance,
            'registrosDetalle': registrosDetalle,
            'usuariosList': usuariosList,

            'subtotalFactura': round(subtotal, 2),
            'totalFactura': round(total, 2),
            'precioTax': precioTax,
            'totalParticipantesMatriculados': totalParticipantesMatriculados,
            'totalHorasServicios': convertir_tiempo(totalHorasServicios),
            
            'serviciosList': Servicio.objects.all(),
            'serviciosFacturados': facturaInstance.proyectosServicios.values_list('servicio',flat=True),
            'registrosPorMes': registrosPorMes,

            'regDetServicio': regDetServicio,
            'sumaTiempoPorRegistro': sumaTiempoPorRegistro
            #'totalParticipantesServidos': totalParticipantesServidos
        }

        html_template = template.render(context)
        
        try:
            pdf = HTML(string=html_template, base_url=request.build_absolute_uri()).write_pdf()
            return HttpResponse(pdf, content_type='application/pdf')
        except Exception as e:
            import traceback
            error_msg = f"Error al generar el PDF: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            return HttpResponse(f"Ha ocurrido un error al generar el PDF. Informe al administrador.\nDetalle: {str(e)}", status=500)

class DescargarExcelRegistros(TemplateView):
    def get(self, request, *args, **kwargs):
        registro_filter = RegistroFilter(request.GET, queryset=Registro.objects.all(), request=self.request)
        
        registrosDetalle = RegistroDetalle.objects.filter(registro__in=registro_filter.qs).order_by('registro', 'fechaHoraInicio')
        
        # Usar DurationField para cálculos de tiempo internos
        diffMiliseg = registrosDetalle.values('registro_id').order_by('registro_id').annotate(timesDif = ExpressionWrapper( (F("fechaHoraFin") - F("fechaHoraInicio")), output_field=django_models.DurationField()) )
        sumaTiempoPorRegistro = diffMiliseg.values('registro_id', 'registro__alumno__nombre', 'registro__alumno__apellidoPaterno', 'registro__alumno__apellidoMaterno', 'timesDif', 'fechaHoraInicio', 'fechaHoraFin', servicio=F('registro__proyecto_servicio__servicio__nombre'), proyecto=F('registro__proyecto_servicio__proyecto__nombre'), escuela=F('registro__alumno__escuela__nombre')).order_by('registro__alumno__nombre')
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el titulo
        ws['B1'] = 'LISTADO REGISTROS POR PARTICIPANTES (HS)'
        #Juntamos las celdas desde la B1 hasta el final, formando una sola celda
        ws.merge_cells('B1:I1')
        #Creamos los encabezados desde la celda B3 hasta la H3
        ws['B3'] = 'REGISTRO ID'
        ws['C3'] = 'PROYECTO'
        ws['D3'] = 'SERVICIO'
        ws['E3'] = 'ESCUELA'
        ws['F3'] = 'ALUMNO'
        ws['G3'] = 'INICIO'
        ws['H3'] = 'FIN'
        ws['I3'] = 'TIEMPO'
        
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 25
        
        for col in range(2,10):
            ws.cell(row=3,column=2).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=3).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=4).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=5).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=6).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=7).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=8).fill = PatternFill (patternType = "solid", start_color = "EFC95D")
            ws.cell(row=3,column=9).fill = PatternFill (patternType = "solid", start_color = "EFC95D")

        row = 4
        #Recorremos el conjunto de registros de alumnos y vamos escribiendo cada uno de los datos en las celdas
        for reg in sumaTiempoPorRegistro:
            ws.cell(row=row,column=2).value = reg['registro_id']
            ws.cell(row=row,column=3).value = reg['proyecto']
            ws.cell(row=row,column=4).value = reg['servicio']
            ws.cell(row=row,column=5).value = reg['escuela']
            ws.cell(row=row,column=6).value = reg['registro__alumno__nombre'] + ' ' + reg['registro__alumno__apellidoPaterno'] + ' ' + reg['registro__alumno__apellidoMaterno']
            ws.cell(row=row,column=7).value = reg['fechaHoraInicio']
            ws.cell(row=row,column=8).value = reg['fechaHoraFin']
            ws.cell(row=row,column=9).value = convertir_tiempo(reg['timesDif'])
            row = row + 1

        # Establecemos el nombre del archivo
        nombre_archivo = 'Default'
        if request.GET['accionDesde'] == 'listarRegistro':
            nombre_archivo ="Listado registros por participante.xlsx"
        elif request.GET['accionDesde'] == 'listarRegistroPorProyecto':
            nombre_archivo ="Listado registros por proyectos.xlsx"

        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response