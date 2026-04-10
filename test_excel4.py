import openpyxl
import json

try:
    wb = openpyxl.load_workbook('C:/s/Sistema/wiao/wioa/WIOA/Tabla - AMSI - Original - 5enero2026.xlsx', data_only=True, read_only=True)
    sheets = wb.sheetnames
    
    grupo_1 = None
    for s in sheets:
        if "Grupo 1" in s:
            grupo_1 = s
            break
            
    res = {"sheets": sheets, "grupo_1_rows": []}
    
    if grupo_1:
        sheet = wb[grupo_1]
        for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
            res["grupo_1_rows"].append([str(c) if c is not None else "" for c in row])
            
    with open('C:/s/Sistema/wiao/wioa/info.json', 'w', encoding='utf-8') as f:
        json.dump(res, f, ensure_ascii=False, indent=2)
except Exception as e:
    import traceback
    traceback.print_exc()
